"""
Multi-format document ingestion.

This module keeps the rest of the application PDF-compatible while allowing
users to upload Word, Excel, CSV, and PDF files.

Design:
  * Every uploaded source is normalized to a PDF for fallback page rendering,
    report compatibility, and best-effort visual coordinate matching.
  * Structured comparison blocks are extracted from the original source format
    when that is more reliable than reading the converted PDF:
      - DOCX: headings, paragraphs, lists, and tables
      - XLSX/XLSM: sheets, rows, and cells
      - CSV: rows and cells
  * Legacy DOC/XLS and parser failures are converted to PDF and then use the
    normal PDF extractor as a fallback.

The frontend can render PDFs through page images + overlays and render
non-PDF sources through the backend's structured native-page endpoint.

LibreOffice is required in the backend container for non-PDF visual conversion.
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

import fitz
from rapidfuzz import fuzz

from .extraction.runner import enrich_blocks
from .models import Block, BlockType, TemplateProfile


SUPPORTED_INPUT_EXTENSIONS = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".tif",
    ".tiff",
    ".bmp",
    ".webp",
    ".docx",
    ".doc",
    ".xlsx",
    ".xlsm",
    ".xlsb",
    ".xls",
    ".csv",
    ".tsv",
}

WORD_EXTENSIONS = {".docx", ".doc"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xlsm", ".xlsb", ".xls", ".csv", ".tsv"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}


def supported_input_extensions() -> list[str]:
    return sorted(SUPPORTED_INPUT_EXTENSIONS)


def source_kind(path: str | Path) -> str:
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return "pdf"
    if ext in IMAGE_EXTENSIONS:
        return "image"
    if ext in WORD_EXTENSIONS:
        return "word"
    if ext in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    return "unknown"


def ensure_supported(path: str | Path) -> None:
    ext = Path(path).suffix.lower()
    if ext not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(
            "Unsupported file type. Supported formats: "
            + ", ".join(supported_input_extensions())
        )


def _hash_content(payload: dict) -> str:
    s = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def _clean(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _slug(value: Any, fallback: str = "section") -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "")).strip("_").lower()
    return text[:70] or fallback


def _header_name(value: Any, index: int) -> str:
    text = _clean(value)
    return text[:90] if text else f"Column {index + 1}"


def _looks_like_identifier(value: Any) -> bool:
    text = _clean(value)
    if not text:
        return False

    low = text.lower()
    if low in {"x", "o", "s", "m", "-", "--", "n/a", "na", "none", "yes", "no", "tbd"}:
        return False
    if "$" in text or "%" in text:
        return False
    if re.fullmatch(r"(?:19|20)\d{2}", text):
        return False
    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text):
        return False

    return bool(
        re.fullmatch(r"[A-Z]{1,10}[- ]?\d{1,12}[A-Z]?", text, re.I)
        or re.fullmatch(r"\d{2,12}[A-Z]?", text, re.I)
        or re.fullmatch(r"[A-Z0-9]{2,12}[A-Z]?", text, re.I)
    )


def _detect_stable_key(row: list[str], header: Optional[list[str]] = None) -> Optional[str]:
    header = header or []
    header_low = [h.lower() for h in header]

    identifier_terms = (
        "id",
        "code",
        "key",
        "number",
        "no",
        "part",
        "item",
        "model",
        "option",
        "order",
        "package",
        "pcv",
        "pcb",
        "sku",
        "ref",
    )

    for i, cell in enumerate(row):
        if not _looks_like_identifier(cell):
            continue
        h = header_low[i] if i < len(header_low) else ""
        if any(term in h for term in identifier_terms):
            return _clean(cell)

    for cell in row:
        if _looks_like_identifier(cell):
            return _clean(cell)

    for cell in row:
        text = _clean(cell)
        if len(text) >= 3:
            return text[:120]

    return None


def _row_payload(header: list[str], row: list[str]) -> dict[str, str]:
    payload: dict[str, str] = {}
    used: set[str] = set()
    max_len = max(len(header), len(row))

    for i in range(max_len):
        key = _header_name(header[i] if i < len(header) else "", i)
        if key in used:
            key = f"{key} {i + 1}"
        used.add(key)
        payload[key] = _clean(row[i]) if i < len(row) else ""

    return payload


def _filled_count(row: list[str]) -> int:
    return sum(1 for cell in row if _clean(cell))


def _looks_like_header_row(row: list[str], body_sample: list[list[str]], n_cols: int) -> bool:
    filled = _filled_count(row)
    if filled == 0:
        return False

    low_text = " ".join(_clean(cell).lower() for cell in row)
    header_terms = (
        "feature",
        "description",
        "item",
        "code",
        "number",
        "no",
        "pcv",
        "pcb",
        "package",
        "model",
        "series",
        "option",
        "value",
        "status",
        "price",
        "date",
        "qty",
        "quantity",
        "size",
        "color",
        "colour",
        "remarks",
        "comments",
    )
    if any(term in low_text for term in header_terms):
        return True

    non_numeric = sum(
        1
        for cell in row
        if _clean(cell) and not re.fullmatch(r"[-+]?[$€£]?\d[\d,]*(?:\.\d+)?%?", _clean(cell))
    )
    body_filled = [_filled_count(sample) for sample in body_sample if _filled_count(sample)]
    avg_body_filled = (sum(body_filled) / len(body_filled)) if body_filled else filled

    if n_cols >= 4 and filled <= max(2, int(n_cols * 0.45)) and non_numeric >= 1:
        return True
    if filled >= 2 and non_numeric >= max(1, filled - 1) and filled <= avg_body_filled + 1:
        return True
    return False


def _merge_header_rows(header_rows: list[list[str]], n_cols: int) -> list[str]:
    merged: list[str] = []
    previous = ""

    for col in range(n_cols):
        parts = []
        last = ""
        for row in header_rows:
            value = _clean(row[col] if col < len(row) else "")
            if not value:
                continue
            if value == last:
                continue
            parts.append(value)
            last = value

        name = " / ".join(parts).strip()
        if not name:
            name = f"Column {col + 1}"
        if name == previous and len(parts) == 1:
            name = f"{name} {col + 1}"
        previous = name
        merged.append(name[:140])

    return merged


def _detect_header_band(normalized_rows: list[list[str]], n_cols: int) -> tuple[list[str], list[list[str]], list[list[str]], int, str]:
    """
    Detect one or more header rows and merge nested headers into stable column
    names. This handles spreadsheets/PDF-derived tables where a group header
    spans several value columns or where the visible header is split over
    multiple rows.
    """
    if not normalized_rows:
        return [], [], [], 0, "empty"

    header_start = 0
    for idx, row in enumerate(normalized_rows[:12]):
        if _filled_count(row) >= max(1, min(2, n_cols)):
            header_start = idx
            break

    body_sample = normalized_rows[header_start + 1 : header_start + 8]
    header_count = 1
    for offset in range(1, min(4, len(normalized_rows) - header_start)):
        row = normalized_rows[header_start + offset]
        next_body = normalized_rows[header_start + offset + 1 : header_start + offset + 6]
        if _looks_like_header_row(row, next_body or body_sample, n_cols):
            header_count += 1
        else:
            break

    header_rows = normalized_rows[header_start : header_start + header_count]
    header = _merge_header_rows(header_rows, n_cols)
    body_rows = normalized_rows[header_start + header_count :]

    if not body_rows:
        body_rows = normalized_rows[header_start + 1 :] if len(normalized_rows) > header_start + 1 else normalized_rows
        header_rows = [normalized_rows[header_start]]
        header = [_header_name(cell, i) for i, cell in enumerate(normalized_rows[header_start])]
        header_count = 1

    return header, body_rows, header_rows, header_start, "nested_header" if header_count > 1 else "single_header"


def _looks_like_layout_table(rows: list[list[str]], n_cols: int) -> bool:
    """
    Word documents often use borderless tables for layout: bilingual text,
    signatures, clauses in two languages, or side-by-side notes. Those should
    render as document content, not as data tables.
    """
    if n_cols < 2 or n_cols > 3 or len(rows) < 2:
        return False

    filled_rows = [row for row in rows if _filled_count(row) >= 1]
    if not filled_rows:
        return False

    long_cell_count = 0
    short_code_count = 0
    numeric_like_count = 0
    total_cells = 0
    headerish_terms = 0

    for row in filled_rows[:12]:
        for cell in row[:n_cols]:
            text = _clean(cell)
            if not text:
                continue
            total_cells += 1
            if len(text) > 45 or len(text.split()) >= 7:
                long_cell_count += 1
            if _looks_like_identifier(text):
                short_code_count += 1
            if re.fullmatch(r"[-+]?[$€£]?\d[\d,]*(?:\.\d+)?%?", text):
                numeric_like_count += 1
            if re.search(r"\b(feature|item|code|pcv|pcb|qty|quantity|price|status|value)\b", text, re.I):
                headerish_terms += 1

    if total_cells == 0:
        return False

    long_ratio = long_cell_count / total_cells
    structured_ratio = (short_code_count + numeric_like_count + headerish_terms) / total_cells
    return long_ratio >= 0.35 and structured_ratio < 0.35


def _row_text(payload: dict[str, str]) -> str:
    parts = []
    for key, value in payload.items():
        if not _clean(value):
            continue
        if re.fullmatch(r"Column \d+", str(key), re.I):
            parts.append(_clean(value))
        else:
            parts.append(f"{key}: {_clean(value)}")
    return " | ".join(parts)


def _text_for_visual_match(block: Block) -> str:
    parts = [block.text or "", block.stable_key or "", block.path or ""]

    if isinstance(block.payload, dict):
        for key, value in block.payload.items():
            key = str(key)
            if key.startswith("__") or key in {"page_width", "page_height", "source_extraction"}:
                continue
            if isinstance(value, list):
                parts.extend(str(v or "") for v in value[:30])
            elif isinstance(value, dict):
                parts.extend(f"{k} {v}" for k, v in list(value.items())[:30])
            else:
                parts.append(str(value or ""))

    return _clean(" ".join(parts))


def _visual_match_score(native: Block, visual: Block) -> float:
    native_text = _text_for_visual_match(native)
    visual_text = _text_for_visual_match(visual)

    if not native_text or not visual_text:
        return 0.0

    token_score = fuzz.token_set_ratio(native_text, visual_text) / 100.0
    partial_score = fuzz.partial_ratio(native_text, visual_text) / 100.0
    ratio_score = fuzz.ratio(native_text, visual_text) / 100.0
    type_bonus = 0.06 if native.block_type == visual.block_type else 0.0

    if native.stable_key and native.stable_key == visual.stable_key:
        type_bonus += 0.12

    return min(1.0, max(token_score, partial_score * 0.92, ratio_score * 0.85) + type_bonus)


def _visual_candidates(native: Block, visual_blocks: list[Block]) -> list[Block]:
    if native.block_type == BlockType.TABLE_ROW:
        preferred = [b for b in visual_blocks if b.block_type == BlockType.TABLE_ROW and b.bbox]
        if preferred:
            return preferred

    if native.block_type == BlockType.TABLE:
        preferred = [b for b in visual_blocks if b.block_type == BlockType.TABLE and b.bbox]
        if preferred:
            return preferred

    if native.block_type in {BlockType.SECTION, BlockType.HEADING}:
        preferred = [
            b for b in visual_blocks
            if b.block_type in {BlockType.SECTION, BlockType.HEADING, BlockType.PARAGRAPH} and b.bbox
        ]
        if preferred:
            return preferred

    return [b for b in visual_blocks if b.bbox]


def _attach_visual_bboxes(native_blocks: list[Block], visual_blocks: list[Block]) -> list[Block]:
    """
    Native Office extraction gives better semantic structure, but it has no PDF
    coordinates. Match native blocks back to the converted PDF blocks so the
    side-by-side viewer can still highlight meaningful regions.
    """
    if not native_blocks or not visual_blocks:
        return native_blocks

    used_visual: set[Any] = set()
    by_type_threshold = {
        BlockType.TABLE_ROW: 0.54,
        BlockType.TABLE: 0.50,
        BlockType.SECTION: 0.58,
        BlockType.HEADING: 0.58,
        BlockType.PARAGRAPH: 0.60,
        BlockType.LIST_ITEM: 0.58,
        BlockType.KV_PAIR: 0.58,
    }

    for block in native_blocks:
        if block.bbox:
            continue

        candidates = _visual_candidates(block, visual_blocks)
        best = None
        best_score = 0.0

        for candidate in candidates:
            if candidate.id in used_visual and block.block_type != BlockType.TABLE:
                continue

            score = _visual_match_score(block, candidate)
            if score > best_score:
                best = candidate
                best_score = score

        threshold = by_type_threshold.get(block.block_type, 0.60)
        if not best or best_score < threshold:
            continue

        block.bbox = best.bbox
        block.page_number = best.page_number
        if isinstance(best.payload, dict):
            block.payload["page_width"] = best.payload.get("page_width", block.payload.get("page_width", 612))
            block.payload["page_height"] = best.payload.get("page_height", block.payload.get("page_height", 792))
        block.payload["visual_match_score"] = round(best_score, 3)
        block.payload["visual_match_source"] = "converted_pdf"

        if block.block_type != BlockType.TABLE:
            used_visual.add(best.id)

    return native_blocks


def _block(
    *,
    block_type: BlockType,
    path: str,
    text: str = "",
    payload: Optional[dict[str, Any]] = None,
    sequence: int = 0,
    page_number: int = 1,
    parent_id: Any = None,
    stable_key: Optional[str] = None,
) -> Block:
    payload = payload or {}
    payload.setdefault("source_extraction", "native")
    payload.setdefault("page_width", 612)
    payload.setdefault("page_height", 792)

    block = Block(
        parent_id=parent_id,
        block_type=block_type,
        path=path,
        page_number=max(1, int(page_number or 1)),
        bbox=None,
        text=text,
        payload=payload,
        sequence=sequence,
        stable_key=stable_key,
    )
    block.content_hash = _hash_content(payload if payload else {"text": text})
    return block


def _page_for_sequence(seq: int, rows_per_page: int = 45) -> int:
    return max(1, int(seq // rows_per_page) + 1)


def _ocr_text_to_blocks(text: str, *, source_path: Path, page_number: int = 1, sequence_start: int = 0) -> list[Block]:
    lines = [_clean(line) for line in str(text or "").splitlines()]
    lines = [line for line in lines if line]
    blocks: list[Block] = []
    seq = sequence_start
    current_section = None
    base_path = "/ocr"

    for line in lines:
        low = line.lower()
        is_heading = (
            seq == sequence_start
            or (len(line) <= 90 and line[:1].isupper() and not re.search(r"[:：]\s*\S", line) and len(line.split()) <= 8)
            or any(term in low for term in ("template", "confirmation", "invoice", "statement", "purchase order"))
        )
        block_type = BlockType.SECTION if is_heading and len(blocks) == 0 else BlockType.PARAGRAPH
        if re.match(r"^\s*[^:：]{2,80}\s*[:：]\s*\S+", line):
            block_type = BlockType.KV_PAIR

        if block_type == BlockType.SECTION:
            path = f"{base_path}/{_slug(line, f'section_{seq}')}"
            block = _block(
                block_type=BlockType.SECTION,
                path=path,
                text=line,
                payload={
                    "text": line,
                    "source_format": source_kind(source_path),
                    "source_extraction": "ocr",
                    "ocr": True,
                },
                sequence=seq,
                page_number=page_number,
            )
            current_section = block
        else:
            parent_path = current_section.path if current_section else base_path
            block = _block(
                parent_id=current_section.id if current_section else None,
                block_type=block_type,
                path=f"{parent_path}/line_{seq}",
                text=line,
                payload={
                    "text": line,
                    "source_format": source_kind(source_path),
                    "source_extraction": "ocr",
                    "ocr": True,
                },
                sequence=seq,
                page_number=page_number,
            )

        blocks.append(block)
        seq += 1

    return blocks


def _extract_image_ocr(source_path: Path) -> list[Block]:
    try:
        from PIL import Image
        import pytesseract
    except Exception:
        return []

    try:
        image = Image.open(str(source_path))
        text = pytesseract.image_to_string(image, lang=os.getenv("TESSERACT_LANG", "eng+ara"))
    except Exception:
        return []

    return _ocr_text_to_blocks(text, source_path=source_path, page_number=1)


def _extract_pdf_ocr(pdf_path: Path, source_path: Path) -> list[Block]:
    try:
        from PIL import Image
        import pytesseract
    except Exception:
        return []

    blocks: list[Block] = []
    seq = 0
    try:
        doc = fitz.open(str(pdf_path))
        for page_idx, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5), alpha=False)
            image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            text = pytesseract.image_to_string(image, lang=os.getenv("TESSERACT_LANG", "eng+ara"))
            page_blocks = _ocr_text_to_blocks(text, source_path=source_path, page_number=page_idx, sequence_start=seq)
            blocks.extend(page_blocks)
            seq += len(page_blocks)
        doc.close()
    except Exception:
        return []

    return blocks


def _safe_copy_upload_name(filename: str, side: str) -> str:
    ext = Path(filename or "").suffix.lower()
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(filename or side).stem).strip("._")
    return f"{side}_{stem or side}{ext or '.bin'}"


def save_upload_to_source(upload_file, work_dir: Path, side: str) -> Path:
    filename = upload_file.filename or f"{side}.pdf"
    path = work_dir / _safe_copy_upload_name(filename, side)
    ensure_supported(path)

    with path.open("wb") as f:
        shutil.copyfileobj(upload_file.file, f)

    return path


def _find_libreoffice() -> Optional[str]:
    configured = os.getenv("LIBREOFFICE_BIN") or os.getenv("SOFFICE_BIN")
    if configured and Path(configured).exists():
        return configured

    for candidate in ("soffice", "libreoffice"):
        found = shutil.which(candidate)
        if found:
            return found

    return None


def normalize_to_pdf(source_path: Path, out_dir: Path) -> Path:
    """
    Return a PDF path for visual rendering.

    PDFs pass through unchanged. Office/CSV files are converted through
    LibreOffice to preserve layout, fonts, tables, and pagination as much as
    possible.
    """
    ensure_supported(source_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    if source_path.suffix.lower() == ".pdf":
        return source_path

    if source_path.suffix.lower() in IMAGE_EXTENSIONS:
        pdf_path = out_dir / f"{source_path.stem}.pdf"
        try:
            img_doc = fitz.open(str(source_path))
            pdf_bytes = img_doc.convert_to_pdf()
            img_doc.close()
            pdf_doc = fitz.open("pdf", pdf_bytes)
            pdf_doc.save(str(pdf_path))
            pdf_doc.close()
        except Exception as exc:
            raise RuntimeError(f"Image conversion to PDF failed: {exc}") from exc

        if not pdf_path.exists():
            raise RuntimeError("Image conversion to PDF failed because no PDF was produced.")
        return pdf_path

    soffice = _find_libreoffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice/soffice is required to convert Word, Excel, or CSV files to PDF. "
            "Install LibreOffice in the backend container or upload PDF files."
        )

    cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        str(out_dir),
        str(source_path),
    ]

    try:
        completed = subprocess.run(
            cmd,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=int(os.getenv("DOCUMENT_CONVERSION_TIMEOUT", "120")),
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f"Document conversion timed out after {exc.timeout} seconds.") from exc

    pdf_path = out_dir / f"{source_path.stem}.pdf"
    if completed.returncode != 0 or not pdf_path.exists():
        raise RuntimeError(
            "Document conversion to PDF failed. "
            f"stdout={completed.stdout[-800:]} stderr={completed.stderr[-800:]}"
        )

    return pdf_path


def _iter_docx_blocks(document):
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    body = document.element.body

    for child in body.iterchildren():
        if child.tag.endswith("}p"):
            yield Paragraph(child, document)
        elif child.tag.endswith("}tbl"):
            yield Table(child, document)


def _extract_docx(source_path: Path) -> list[Block]:
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except Exception:
        return []

    document = Document(str(source_path))
    blocks: list[Block] = []
    seq = 0
    path_stack = ["document"]
    current_section = None

    for item in _iter_docx_blocks(document):
        if isinstance(item, Paragraph):
            text = _clean(item.text)
            if not text:
                continue

            style_name = _clean(getattr(item.style, "name", ""))
            is_heading = style_name.lower().startswith("heading")

            if is_heading:
                level_match = re.search(r"(\d+)", style_name)
                level = int(level_match.group(1)) if level_match else 1
                level = max(1, min(6, level))
                path_stack = path_stack[:level] + [_slug(text, f"heading_{seq}")]
                path = "/" + "/".join(path_stack)
                block = _block(
                    block_type=BlockType.SECTION,
                    path=path,
                    text=text,
                    payload={
                        "heading": text,
                        "style": style_name,
                        "source_format": "docx",
                    },
                    sequence=seq,
                    page_number=_page_for_sequence(seq),
                    parent_id=current_section.id if current_section and level > 1 else None,
                )
                blocks.append(block)
                current_section = block
                seq += 1
                continue

            block_type = BlockType.LIST_ITEM if style_name.lower().startswith("list") or re.match(r"^[-*•]\s+", text) else BlockType.PARAGRAPH
            base_path = current_section.path if current_section else "/document"
            block = _block(
                parent_id=current_section.id if current_section else None,
                block_type=block_type,
                path=f"{base_path}/p_{seq}",
                text=text,
                payload={
                    "text": text,
                    "style": style_name,
                    "source_format": "docx",
                },
                sequence=seq,
                page_number=_page_for_sequence(seq),
            )
            blocks.append(block)
            seq += 1
            continue

        if isinstance(item, Table):
            rows = []
            for raw_row in item.rows:
                rows.append([_clean(cell.text) for cell in raw_row.cells])

            rows = [row for row in rows if any(_clean(cell) for cell in row)]
            if not rows:
                continue

            n_cols = max(len(row) for row in rows)
            normalized_rows = [row + [""] * (n_cols - len(row)) for row in rows]

            if _looks_like_layout_table(normalized_rows, n_cols):
                base_path = current_section.path if current_section else "/document"
                for ri, row in enumerate(normalized_rows):
                    row_text = " / ".join(_clean(cell) for cell in row if _clean(cell))
                    if not row_text:
                        continue
                    block = _block(
                        parent_id=current_section.id if current_section else None,
                        block_type=BlockType.PARAGRAPH,
                        path=f"{base_path}/layout_{seq}",
                        text=row_text,
                        payload={
                            "text": row_text,
                            "source_format": "docx",
                            "layout_table": True,
                            "layout_columns": [_clean(cell) for cell in row if _clean(cell)],
                        },
                        sequence=seq,
                        page_number=_page_for_sequence(seq),
                    )
                    blocks.append(block)
                    seq += 1
                continue

            header, body_rows, header_rows, header_index, header_strategy = _detect_header_band(normalized_rows, n_cols)

            base_path = current_section.path if current_section else "/document"
            table_title = _clean(current_section.text if current_section else "") or f"Table {len([b for b in blocks if b.block_type == BlockType.TABLE]) + 1}"
            table_payload = {
                "header": header,
                "header_rows": header_rows,
                "header_row_count": len(header_rows),
                "header_index": header_index,
                "header_strategy": header_strategy,
                "rows": body_rows,
                "spans_pages": [_page_for_sequence(seq)],
                "table_title": table_title,
                "table_context": base_path.replace("_", " ").strip("/"),
                "source_format": "docx",
            }
            table_block = _block(
                parent_id=current_section.id if current_section else None,
                block_type=BlockType.TABLE,
                path=f"{base_path}/table_{seq}",
                text=table_title,
                payload=table_payload,
                sequence=seq,
                page_number=_page_for_sequence(seq),
            )
            blocks.append(table_block)
            seq += 1

            for ri, row in enumerate(body_rows):
                payload = _row_payload(header, row)
                payload.update(
                    {
                        "__row_index__": ri,
                        "__table_title__": table_title,
                        "__pages__": [_page_for_sequence(seq)],
                        "source_format": "docx",
                    }
                )
                text = _row_text(payload)
                block = _block(
                    parent_id=table_block.id,
                    block_type=BlockType.TABLE_ROW,
                    path=f"{table_block.path}/row_{ri}",
                    text=text,
                    payload=payload,
                    sequence=seq,
                    page_number=_page_for_sequence(seq),
                    stable_key=_detect_stable_key(row, header),
                )
                blocks.append(block)
                seq += 1

    return blocks


def _sheet_rows_from_openpyxl(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(source_path), data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = []
            for raw_row in sheet.iter_rows(values_only=True):
                row = [_clean(value) for value in raw_row]
                if any(row):
                    rows.append(row)
            yield sheet.title, rows
    finally:
        workbook.close()


def _sheet_rows_from_xls(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except Exception:
        return []

    workbook = xlrd.open_workbook(str(source_path))
    out = []
    for sheet in workbook.sheets():
        rows = []
        for ri in range(sheet.nrows):
            row = [_clean(sheet.cell_value(ri, ci)) for ci in range(sheet.ncols)]
            if any(row):
                rows.append(row)
        out.append((sheet.name, rows))
    return out


def _sheet_rows_from_xlsb(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    try:
        from pyxlsb import open_workbook
    except Exception:
        return []

    out = []
    with open_workbook(str(source_path)) as workbook:
        for sheet_name in workbook.sheets:
            rows = []
            with workbook.get_sheet(sheet_name) as sheet:
                for raw_row in sheet.rows():
                    row = [_clean(cell.v if cell is not None else "") for cell in raw_row]
                    if any(row):
                        rows.append(row)
            out.append((sheet_name, rows))
    return out


def _sheet_rows_from_csv(source_path: Path) -> Iterable[tuple[str, list[list[str]]]]:
    delimiter = "\t" if source_path.suffix.lower() == ".tsv" else ","
    rows = []
    with source_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            clean = [_clean(cell) for cell in row]
            if any(clean):
                rows.append(clean)
    return [(source_path.stem, rows)]


def _extract_spreadsheet(source_path: Path) -> list[Block]:
    ext = source_path.suffix.lower()

    try:
        if ext in {".xlsx", ".xlsm"}:
            sheets = list(_sheet_rows_from_openpyxl(source_path))
        elif ext == ".xlsb":
            sheets = list(_sheet_rows_from_xlsb(source_path))
        elif ext == ".xls":
            sheets = list(_sheet_rows_from_xls(source_path))
        else:
            sheets = list(_sheet_rows_from_csv(source_path))
    except Exception:
        return []

    blocks: list[Block] = []
    seq = 0

    for sheet_name, rows in sheets:
        rows = [row for row in rows if any(_clean(cell) for cell in row)]
        if not rows:
            continue

        n_cols = max(len(row) for row in rows)
        normalized_rows = [row + [""] * (n_cols - len(row)) for row in rows]

        header, body_rows, header_rows, header_index, header_strategy = _detect_header_band(normalized_rows, n_cols)

        sheet_slug = _slug(sheet_name, f"sheet_{seq}")
        section = _block(
            block_type=BlockType.SECTION,
            path=f"/{sheet_slug}",
            text=sheet_name,
            payload={
                "heading": sheet_name,
                "source_format": ext.lstrip("."),
                "sheet_name": sheet_name,
            },
            sequence=seq,
            page_number=_page_for_sequence(seq, 55),
        )
        blocks.append(section)
        seq += 1

        table_payload = {
            "header": header,
            "header_rows": header_rows,
            "header_row_count": len(header_rows),
            "header_index": header_index,
            "header_strategy": header_strategy,
            "rows": body_rows,
            "spans_pages": [_page_for_sequence(seq, 55)],
            "table_title": sheet_name,
            "table_context": f"Sheet: {sheet_name}",
            "source_format": ext.lstrip("."),
            "sheet_name": sheet_name,
        }
        table = _block(
            parent_id=section.id,
            block_type=BlockType.TABLE,
            path=f"/{sheet_slug}/table_{seq}",
            text=sheet_name,
            payload=table_payload,
            sequence=seq,
            page_number=_page_for_sequence(seq, 55),
        )
        blocks.append(table)
        seq += 1

        for ri, row in enumerate(body_rows):
            payload = _row_payload(header, row)
            payload.update(
                {
                    "__row_index__": ri,
                    "__table_title__": sheet_name,
                    "__pages__": [_page_for_sequence(seq, 55)],
                    "source_format": ext.lstrip("."),
                    "sheet_name": sheet_name,
                }
            )
            text = _row_text(payload)
            block = _block(
                parent_id=table.id,
                block_type=BlockType.TABLE_ROW,
                path=f"{table.path}/row_{ri}",
                text=text,
                payload=payload,
                sequence=seq,
                page_number=_page_for_sequence(seq, 55),
                stable_key=_detect_stable_key(row, header),
            )
            blocks.append(block)
            seq += 1

    return blocks


def extract_blocks_from_source(
    source_path: Path,
    pdf_path: Path,
    pdf_extractor: Callable[[str], list[Block]],
    profile: Optional[TemplateProfile] = None,
) -> list[Block]:
    """
    Extract structured blocks from the best available source.

    profile is accepted for API compatibility; native Office extraction is
    intentionally generic and template-free in this first release.
    """
    ext = source_path.suffix.lower()
    source_format = source_kind(source_path)
    document_label = source_path.stem

    if ext in IMAGE_EXTENSIONS:
        blocks = pdf_extractor(str(pdf_path))
        extracted_chars = len(re.sub(r"\s+", "", " ".join(block.text or "" for block in blocks)))
        if extracted_chars < 25:
            ocr_blocks = _extract_image_ocr(source_path)
            if ocr_blocks:
                blocks = ocr_blocks
        return enrich_blocks(
            blocks,
            source_path=source_path,
            source_format=source_format,
            document_label=document_label,
        )

    if ext == ".pdf":
        blocks = pdf_extractor(str(pdf_path))
        extracted_chars = len(re.sub(r"\s+", "", " ".join(block.text or "" for block in blocks)))
        if extracted_chars < 25:
            ocr_blocks = _extract_pdf_ocr(pdf_path, source_path)
            if ocr_blocks:
                blocks = ocr_blocks
        return enrich_blocks(
            blocks,
            source_path=source_path,
            source_format=source_format,
            document_label=document_label,
        )

    blocks: list[Block] = []
    if ext == ".docx":
        blocks = _extract_docx(source_path)
    elif ext in SPREADSHEET_EXTENSIONS:
        blocks = _extract_spreadsheet(source_path)

    if blocks:
        visual_blocks = pdf_extractor(str(pdf_path))
        blocks = _attach_visual_bboxes(blocks, visual_blocks)
        return enrich_blocks(
            blocks,
            source_path=source_path,
            source_format=source_format,
            document_label=document_label,
        )

    # DOC/legacy XLS or parser failure: use the converted PDF as a safe fallback.
    blocks = pdf_extractor(str(pdf_path))
    return enrich_blocks(
        blocks,
        source_path=source_path,
        source_format=source_format,
        document_label=document_label,
    )


def coverage_for_source(source_path: Path, pdf_path: Path, blocks: list[Block], pdf_coverage: Callable[[str, list[Block]], float]) -> float:
    if source_path.suffix.lower() == ".pdf":
        return pdf_coverage(str(pdf_path), blocks)

    extracted = len(re.sub(r"\s+", "", " ".join(block.text or "" for block in blocks)))
    if extracted > 0:
        return 100.0
    return 0.0
